'''
    基于excel驱动自动化测试操作行为
'''
import openpyxl
import traceback

from class18_web_keys.web_keys import WebKeys

# 函数定义：接受一个名为 data 的参数。这个参数通常是一个字符串，也可能为 None。
def arguments(data):
    temp_data = {}  # 初始化一个空的字典，用于存放解析后得到的键值对。
    # 参数解析：要么为None，要么为str
    if data:    # 判断为None则不处理（表格中C1、C2、C11是None，无需进行处理）
        str_temp = data.split(';')  # 将str基于分号分割，生成list。（表格中不同的参数使用的是;进行分割）。
        """
            调用字符串的 split 方法，以分号 ';' 为分隔符，将整个字符串切分成多个子字符串。
            返回一个列表 (list)。
            例如：
                如果 data = 'by=id;value=chat-textarea;txt=SAD'
                那么 str_temp = ['by=id', 'value=chat-textarea', 'txt=SAD']
        """
        for temp in str_temp:   # 进行二次分割：对by=id、value=chat-textarea、txt=SAD进行解析
            """
                这是一个 for 循环，遍历上面生成的列表 str_temp 中的每一个元素 (sub-string)。
                第一次循环，temp = 'by=id'
                第二次循环，temp = 'value=chat-textarea'
                第三次循环，temp = 'txt=SAD'
            """
            t = temp.split('=', 1)  # 基于=进行二次分割，参数 1 表示最多只分割一次，结果为一个包含两个元素的列表。
            """
                # 对当前循环的子字符串 temp，再次调用 split 方法。
                # 这次以等号 '=' 为分隔符。
                # 参数 1 表示最多只分割一次 (maxsplit=1)，这样做是为了防止 value 的值本身包含等号时被错误地再次分割。
                    # 例如，如果 temp = 'url=http://example.com?param=value'，split('=') 会变成 ['url', 'http:', '//example.com?param', 'value']
                    # 但 split('=', 1) 会变成 ['url', 'http://example.com?param=value']，这才是正确的。
                # 返回一个包含两个元素的列表 t。
                # 第一次循环：temp='by=id', t=['by', 'id']
                # 第二次循环：temp='value=chat-textarea', t=['value', 'chat-textarea']
                # 第三次循环：temp='txt=SAD', t=['txt', 'SAD']
            """

            temp_data[t[0]] = t[1]  # 二次分割后的列表t中，t[0]为key，t[1]为value，并赋值到字典中
            """
                # 对当前循环的子字符串 temp，再次调用 split 方法。
                # 这次以等号 '=' 为分隔符。
                # 参数 1 表示最多只分割一次 (maxsplit=1)。
                # 这样做是为了防止 value 的值本身包含等号时被错误地再次分割。
                # 例如，如果 temp = 'url=http://example.com?param=value'，split('=') 会变成 ['url', 'http:', '//example.com?param', 'value']
                # 但 split('=', 1) 会变成 ['url', 'http://example.com?param=value']，这才是正确的。
                # 返回一个包含两个元素的列表 t。
                # 第一次循环：temp='by=id', t=['by', 'id']
                # 第二次循环：temp='value=chat-textarea', t=['value', 'chat-textarea']
                # 第三次循环：temp='txt=SAD', t=['txt', 'SAD']
            """
            """
                例如：browser_type=Chrome
                解析后变为['browser_type', 'Chrome']
                赋值后变为temp_data['browser_type'] = 'Chrome'
                也就意味着temp_data = {
                    'browser_type': 'Chrome'
                }
            """
    return temp_data    # 函数结束时，返回解析得到的字典。
    # 无论 if 条件是否执行，函数最后都会返回 temp_data。
    # 如果 data 为 None 或空，if 块不执行，temp_data 保持为空字典 {}，就返回 {}。
    # 如果 data 有效，temp_data 就包含了所有解析出的键值对，返回这个填充好的字典。


# 读取excel文件
excel = openpyxl.load_workbook('./demo.xlsx')
sheet1 = excel['Sheet1']    # 加载表Sheet1（就是下面那一行显示有哪些表）
sheet2 = excel['Sheet2']    # 加载表Sheet2

# 获取用例内容
for value in sheet1.values: # 遍历表格；value代表每一行的内容；value[i]代表元组value中的第几个值。例如：value=('编号', '操作步骤', '参数', '描述', '预期结果', '实际结果'), value[0]='编号', value[1]='操作步骤'....
    # 基于用例编号，来判断正文。value[0]就是第一列的内容，以此类推。
    if type(value[0]) is int:   # 判断【第1列】是否为数字，如果是数字，就执行测试步骤
        # print(value)  # 打印该行的内容

        # 后续可以用logging替换下面这一行的代码
        print('当前正在执行操作：' + value[3])   # 打印：当前正在执行操作+第4列的“描述”

        # 解析测试参数：将用例中的str变为操作方法的操作格式
        data = arguments(value[2])  # 将第3列的“参数”进行解析：基于**kwargs定值不定长传参逻辑，实现对参数的二次处理。

        """
            操作步骤其实就分为以下类型：
                1. 实例化
                2. 常规操作行为
                3. 断言：因为需要对excel文件进行写入。
        """
        if value[1] == 'open_browser':  # 实例化
            wk = WebKeys(**data)    # 调用class18_web_keys/web_keys.py的WebKeys类，创建wk对象，并赋值为 WebKeys 类的一个实例。
        elif 'assert' in value[1]:  # 断言，判断操作指令是否包含 "assert"（如 assert_text, assert_title）
            # 通过与否是基于断言是否报错来决定的
            try:
                getattr(wk, value[1])(**data, expected=value[4])    # 传入预期结果数据。wk 已经在上一次循环中被定义了（如果上一行是 open_browser），当遇到 assert 或其他操作时，wk 变量已经存在，可以直接使用
                """
                    getattr(wk, value[1])(**data, expected=value[4])
                    # 这行代码的执行前提是：wk 已经在之前的循环中通过 open_browser 操作被定义了
                    # 它的作用是：
                        # 1. getattr(wk, "assert_text") -> 获取 wk 对象的 assert_text 方法，也就是class18_web_keys/web_keys.py的WebKeys类中的assert_text方法
                        # 2. (**data, expected=value[4]) -> 将参数传递给该方法
                        # 3. 最终执行 wk.assert_text(by='...', value='...', expected='...')
                """
                """
                    getattr(wk, value[1])：反射机制，动态获取wk对象的指定方法（如 assert_text）
                    (**data, expected=value[4])：
                        **data：将参数字典解包为关键字参数（如 by='xpath', value='...'）
                        expected=value[4]：value[4] 是 Excel 中第5列的值（预期结果），作为 expected 参数传入
                        例如：wk.assert_text(by='xpath', value='...', expected='生词本')
                """
                sheet1.cell(row=value[0] + 3, column=6).value = "PASS"  # 断言通过，在excel中的“实际结果”写入pass # row：编号+特定的数，得到excel中的行数；而列是固定的
            except:
                traceback.print_exc()
                sheet1.cell(row=value[0] + 3, column=6).value = "Failed" # 断言不通过，在excel中的“实际结果”Failed
            finally:
                excel.save('./demo.xlsx')
        else:
            """
                else：处理非初始化、非断言的操作（如 click, input, wait 等）
                getattr(wk, value[1])：动态获取操作方法（如 click, input）
                (**data)：将参数字典解包为关键字参数传递给方法
                例如：wk.click(by='id', value='submit-btn')
            """
            getattr(wk, value[1])(**data)   # 操作行为基于反射机制实现


















